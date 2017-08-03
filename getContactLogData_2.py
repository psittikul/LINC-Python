import openpyxl
wb = openpyxl.load_workbook('LINC Demographics for LVAIC_2.xlsx')
contactLogSheet = wb.get_sheet_by_name('Contact Log Info')







# Cycle through all the rows of our contact log data
for row in range(2, 1320):
    # Go through each cell in the row to get clientID, time, and service
    for cell in range(1, 5):
        clientID = contactLogSheet['A'+str(row)].value
        time = contactLogSheet['C' + str(row)].value
        service = contactLogSheet['D' + str(row)].value
        category = contactLogSheet['E' + str(row)].value
    # Now that we've gotten the necessary information from this row,
    # take our stored values, find the corresponding row in the Ref tab,
    # copy the rest of the necessary information for this client,
    # and create a new row in our data set sheet with the complete information set
    refSheet = wb.get_sheet_by_name('Ref')
    # Going through all the rows in our Ref sheet
    for refRow in range(2, refSheet.max_row + 1):
        # If this row is the client corresponding to the data we got from the current contact log data row,
        # fill in the rest of the necessary data, grabbing what info you need from the ref sheet
        if refSheet['A'+str(refRow)].value == clientID:
            refSheet = wb.get_sheet_by_name('Ref')
            customer = refSheet['B'+str(refRow)].value
            intakeDate = refSheet['C'+str(refRow)].value
            serviceName = refSheet['D'+str(refRow)].value
            clientAge = refSheet['E'+str(refRow)].value
            position = refSheet['F'+str(refRow)].value
            relocating = refSheet['G'+str(refRow)].value
            ethnicity = refSheet['H'+str(refRow)].value
            groups = refSheet['I'+str(refRow)].value
            familyComp = refSheet['J'+str(refRow)].value
            rentOwn = refSheet['K'+str(refRow)].value


            # Function to make the correct service row when the client has multiple services
            def multipleServices(service, serviceName, time):
                if service == 'CT & DC':
                    time = time / 2
                    testSheet = wb.get_sheet_by_name('Test')
                    newRow = str(testSheet.max_row + 1)
                    testSheet['A' + newRow].value = clientID
                    testSheet['B' + newRow].value = customer
                    testSheet['C' + newRow].value = intakeDate
                    testSheet['D' + newRow].value = serviceName
                    testSheet['E' + newRow].value = clientAge
                    testSheet['F' + newRow].value = position
                    testSheet['G' + newRow].value = relocating
                    testSheet['H' + newRow].value = ethnicity
                    testSheet['I' + newRow].value = groups
                    testSheet['J' + newRow].value = familyComp
                    testSheet['K' + newRow].value = rentOwn
                    testSheet['M' + newRow].value = time
                    testSheet['N' + newRow].value = category
                    wb.save('LINC Demographics for LVAIC_2.xlsx')
                    newRow = str(testSheet.max_row + 1)
                    refSheet = wb.get_sheet_by_name('Ref')
                    customer_2 = refSheet['B' + str(refRow + 1)].value
                    intakeDate_2 = refSheet['C' + str(refRow + 1)].value
                    serviceName_2 = refSheet['D' + str(refRow + 1)].value
                    clientAge_2 = refSheet['E' + str(refRow + 1)].value
                    position_2 = refSheet['F' + str(refRow + 1)].value
                    relocating_2 = refSheet['G' + str(refRow + 1)].value
                    ethnicity_2 = refSheet['H' + str(refRow + 1)].value
                    groups_2 = refSheet['I' + str(refRow + 1)].value
                    familyComp_2 = refSheet['J' + str(refRow + 1)].value
                    rentOwn_2 = refSheet['K' + str(refRow + 1)].value
                    testSheet = wb.get_sheet_by_name('Test')
                    testSheet['A' + newRow].value = clientID
                    testSheet['B' + newRow].value = customer_2
                    testSheet['C' + newRow].value = intakeDate_2
                    testSheet['D' + newRow].value = serviceName_2
                    testSheet['E' + newRow].value = clientAge_2
                    testSheet['F' + newRow].value = position_2
                    testSheet['G' + newRow].value = relocating_2
                    testSheet['H' + newRow].value = ethnicity_2
                    testSheet['I' + newRow].value = groups_2
                    testSheet['J' + newRow].value = familyComp_2
                    testSheet['K' + newRow].value = rentOwn_2
                    testSheet['M' + newRow].value = time
                    testSheet['N' + newRow].value = category
                elif service == 'CT':
                    if 'CT' in serviceName or 'Community Transition' in serviceName:
                        testSheet = wb.get_sheet_by_name('Test')
                        newRow = str(testSheet.max_row + 1)
                        testSheet['A' + newRow].value = clientID
                        testSheet['B' + newRow].value = customer
                        testSheet['C' + newRow].value = intakeDate
                        testSheet['D' + newRow].value = serviceName
                        testSheet['E' + newRow].value = clientAge
                        testSheet['F' + newRow].value = position
                        testSheet['G' + newRow].value = relocating
                        testSheet['H' + newRow].value = ethnicity
                        testSheet['I' + newRow].value = groups
                        testSheet['J' + newRow].value = familyComp
                        testSheet['K' + newRow].value = rentOwn
                        testSheet['M' + newRow].value = time
                        testSheet['N' + newRow].value = category
                        wb.save('LINC Demographics for LVAIC_2.xlsx')
                    # If this current refRow is not the correct service, check the rows above and below for the correct service row
                    else:
                        refSheet = wb.get_sheet_by_name('Ref')
                        if 'CT' in refSheet['D' + str(refRow - 1)].value or 'Community Transition' in refSheet[
                                    'D' + str(refRow - 1)].value:
                            checkCustomer = refSheet['B' + str(refRow - 1)].value
                            checkIntakeDate = refSheet['C' + str(refRow - 1)].value
                            checkServiceName = refSheet['D' + str(refRow - 1)].value
                            checkClientAge = refSheet['E' + str(refRow - 1)].value
                            checkPosition = refSheet['F' + str(refRow - 1)].value
                            checkRelocating = refSheet['G' + str(refRow - 1)].value
                            checkEthnicity = refSheet['H' + str(refRow - 1)].value
                            checkGroups = refSheet['I' + str(refRow - 1)].value
                            checkFamilyComp = refSheet['J' + str(refRow - 1)].value
                            checkRentOwn = refSheet['K' + str(refRow - 1)].value
                            testSheet = wb.get_sheet_by_name('Test')
                            newRow = str(testSheet.max_row + 1)
                            testSheet['A' + newRow].value = clientID
                            testSheet['B' + newRow].value = checkCustomer
                            testSheet['C' + newRow].value = checkIntakeDate
                            testSheet['D' + newRow].value = checkServiceName
                            testSheet['E' + newRow].value = checkClientAge
                            testSheet['F' + newRow].value = checkPosition
                            testSheet['G' + newRow].value = checkRelocating
                            testSheet['H' + newRow].value = checkEthnicity
                            testSheet['I' + newRow].value = checkGroups
                            testSheet['J' + newRow].value = checkFamilyComp
                            testSheet['K' + newRow].value = checkRentOwn
                            testSheet['M' + newRow].value = time
                            testSheet['N' + newRow].value = category
                            wb.save('LINC Demographics for LVAIC_2.xlsx')
                        elif 'CT' in refSheet['D' + str(refRow + 1)].value or 'Community Transition' in refSheet[
                                    'D' + str(refRow + 1)].value:
                            checkCustomer = refSheet['B' + str(refRow + 1)].value
                            checkIntakeDate = refSheet['C' + str(refRow + 1)].value
                            checkServiceName = refSheet['D' + str(refRow + 1)].value
                            checkClientAge = refSheet['E' + str(refRow + 1)].value
                            checkPosition = refSheet['F' + str(refRow + 1)].value
                            checkRelocating = refSheet['G' + str(refRow + 1)].value
                            checkEthnicity = refSheet['H' + str(refRow + 1)].value
                            checkGroups = refSheet['I' + str(refRow + 1)].value
                            checkFamilyComp = refSheet['J' + str(refRow + 1)].value
                            checkRentOwn = refSheet['K' + str(refRow + 1)].value
                            testSheet = wb.get_sheet_by_name('Test')
                            newRow = str(testSheet.max_row + 1)
                            testSheet['A' + newRow].value = clientID
                            testSheet['B' + newRow].value = checkCustomer
                            testSheet['C' + newRow].value = checkIntakeDate
                            testSheet['D' + newRow].value = checkServiceName
                            testSheet['E' + newRow].value = checkClientAge
                            testSheet['F' + newRow].value = checkPosition
                            testSheet['G' + newRow].value = checkRelocating
                            testSheet['H' + newRow].value = checkEthnicity
                            testSheet['I' + newRow].value = checkGroups
                            testSheet['J' + newRow].value = checkFamilyComp
                            testSheet['K' + newRow].value = checkRentOwn
                            testSheet['M' + newRow].value = time
                            testSheet['N' + newRow].value = category
                            wb.save('LINC Demographics for LVAIC_2.xlsx')
                elif service == 'DC':
                    if 'DC' in serviceName or 'Dual Career' in serviceName:
                        testSheet = wb.get_sheet_by_name('Test')
                        newRow = str(testSheet.max_row + 1)
                        testSheet['A' + newRow].value = clientID
                        testSheet['B' + newRow].value = customer
                        testSheet['C' + newRow].value = intakeDate
                        testSheet['D' + newRow].value = serviceName
                        testSheet['E' + newRow].value = clientAge
                        testSheet['F' + newRow].value = position
                        testSheet['G' + newRow].value = relocating
                        testSheet['H' + newRow].value = ethnicity
                        testSheet['I' + newRow].value = groups
                        testSheet['J' + newRow].value = familyComp
                        testSheet['K' + newRow].value = rentOwn
                        testSheet['M' + newRow].value = time
                        testSheet['N' + newRow].value = category
                        wb.save('LINC Demographics for LVAIC_2.xlsx')
                    # If this current refRow is not the correct service, check the rows above and below for the correct service row
                    else:
                        refSheet = wb.get_sheet_by_name('Ref')
                        if 'DC' in refSheet['D' + str(refRow - 1)].value or 'Dual Career' in refSheet[
                                    'D' + str(refRow - 1)].value:
                            checkCustomer = refSheet['B' + str(refRow - 1)].value
                            checkIntakeDate = refSheet['C' + str(refRow - 1)].value
                            checkServiceName = refSheet['D' + str(refRow - 1)].value
                            checkClientAge = refSheet['E' + str(refRow - 1)].value
                            checkPosition = refSheet['F' + str(refRow - 1)].value
                            checkRelocating = refSheet['G' + str(refRow - 1)].value
                            checkEthnicity = refSheet['H' + str(refRow - 1)].value
                            checkGroups = refSheet['I' + str(refRow - 1)].value
                            checkFamilyComp = refSheet['J' + str(refRow - 1)].value
                            checkRentOwn = refSheet['K' + str(refRow - 1)].value
                            testSheet = wb.get_sheet_by_name('Test')
                            newRow = str(testSheet.max_row + 1)
                            testSheet['A' + newRow].value = clientID
                            testSheet['B' + newRow].value = checkCustomer
                            testSheet['C' + newRow].value = checkIntakeDate
                            testSheet['D' + newRow].value = checkServiceName
                            testSheet['E' + newRow].value = checkClientAge
                            testSheet['F' + newRow].value = checkPosition
                            testSheet['G' + newRow].value = checkRelocating
                            testSheet['H' + newRow].value = checkEthnicity
                            testSheet['I' + newRow].value = checkGroups
                            testSheet['J' + newRow].value = checkFamilyComp
                            testSheet['K' + newRow].value = checkRentOwn
                            testSheet['M' + newRow].value = time
                            testSheet['N' + newRow].value = category
                            wb.save('LINC Demographics for LVAIC_2.xlsx')
                        elif 'DC' in refSheet['D' + str(refRow + 1)].value or 'Dual Career' in refSheet[
                                    'D' + str(refRow + 1)].value:
                            checkCustomer = refSheet['B' + str(refRow + 1)].value
                            checkIntakeDate = refSheet['C' + str(refRow + 1)].value
                            checkServiceName = refSheet['D' + str(refRow + 1)].value
                            checkClientAge = refSheet['E' + str(refRow + 1)].value
                            checkPosition = refSheet['F' + str(refRow + 1)].value
                            checkRelocating = refSheet['G' + str(refRow + 1)].value
                            checkEthnicity = refSheet['H' + str(refRow + 1)].value
                            checkGroups = refSheet['I' + str(refRow + 1)].value
                            checkFamilyComp = refSheet['J' + str(refRow + 1)].value
                            checkRentOwn = refSheet['K' + str(refRow + 1)].value
                            testSheet = wb.get_sheet_by_name('Test')
                            newRow = str(testSheet.max_row + 1)
                            testSheet['A' + newRow].value = clientID
                            testSheet['B' + newRow].value = checkCustomer
                            testSheet['C' + newRow].value = checkIntakeDate
                            testSheet['D' + newRow].value = checkServiceName
                            testSheet['E' + newRow].value = checkClientAge
                            testSheet['F' + newRow].value = checkPosition
                            testSheet['G' + newRow].value = checkRelocating
                            testSheet['H' + newRow].value = checkEthnicity
                            testSheet['I' + newRow].value = checkGroups
                            testSheet['J' + newRow].value = checkFamilyComp
                            testSheet['K' + newRow].value = checkRentOwn
                            testSheet['M' + newRow].value = time
                            testSheet['N' + newRow].value = category
                            wb.save('LINC Demographics for LVAIC_2.xlsx')
                else:
                    return


            # Function to create a new row and populate its cells
            def newRow(serviceName):
                # Open the test sheet and copy over the values we know we can fill in right away,
                # and get the row number of the test sheet we will be inserting data into
                testSheet = wb.get_sheet_by_name('Test')
                newRow = str(testSheet.max_row + 1)
                testSheet['A' + newRow].value = clientID
                testSheet['B' + newRow].value = customer
                testSheet['C' + newRow].value = intakeDate
                testSheet['D' + newRow].value = serviceName
                testSheet['E' + newRow].value = clientAge
                testSheet['F' + newRow].value = position
                testSheet['G' + newRow].value = relocating
                testSheet['H' + newRow].value = ethnicity
                testSheet['I' + newRow].value = groups
                testSheet['J' + newRow].value = familyComp
                testSheet['K' + newRow].value = rentOwn
                testSheet['M' + newRow].value = time
                testSheet['N' + newRow].value = category
                wb.save('LINC Demographics for LVAIC_2.xlsx')
            # In cases where the client is receiving multiple services, we need to split up any CT & DC entries between the
            # two rows, or we need to make sure CT time is placed in the CT row and DC time in the DC row
            if 'Yes' in refSheet['O'+str(refRow)].value:
                multipleServices(service, serviceName, time)
            else:
                newRow(serviceName)
        refSheet = wb.get_sheet_by_name('Ref')
    wb.save('LINC Demographics for LVAIC_2.xlsx')
wb.save('LINC Demographics for LVAIC_2.xlsx')