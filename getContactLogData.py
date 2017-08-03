import openpyxl
wb = openpyxl.load_workbook('LINC Demographics for LVAIC.xlsx')
print(type(wb))
contactLogSheet = wb.get_sheet_by_name('Contact Log Info')
print(contactLogSheet['A2'].value)

# Cycle through all the rows of our contact log data
for row in range(2, 1320):
    print('Row #: ' + str(row))
    # Go through each cell in the row to get clientID, time, and service
    for cell in range(1, 5):
        clientID = contactLogSheet['A'+str(row)].value
        time = contactLogSheet['C' + str(row)].value
        service = contactLogSheet['D' + str(row)].value
        # Divide the time by 2 if the service is CT & DC
        if (service == 'CT & DC'):
            time = time/2
    # Now that we've gotten the necessary information from this row,
    # take our stored values, find the corresponding row in the Ref tab,
    # copy the rest of the necessary information for this client,
    # and create a new row in our data set sheet with the complete information set
    refSheet = wb.get_sheet_by_name('Ref')
    for refRow in range(2, refSheet.max_row + 1):
        if refSheet['A'+str(refRow)].value == clientID:
            print('Client #: ' + str(refSheet['A'+str(refRow)].value), refSheet['D'+str(refRow)].value)
            if service == refSheet['D'+str(row)].value:
                print('This client has the same service as this row')
                testSheet = wb.get_sheet_by_name('Test')
                testSheet['A'+str(testSheet.max_row+1)].value = clientID
                testSheet['M'+str(testSheet.max_row+1)].value = time
                wb.save('LINC Demographics for LVAIC.xlsx')
        else:
            print('This is not the correct client')
    print('---- END OF ROW ----')
wb.save('LINC Demographics for LVAIC.xlsx')